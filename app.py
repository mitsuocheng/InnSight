import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from openpyxl import load_workbook
import io
from datetime import date, datetime

st.set_page_config(page_title="InnSight", layout="centered")

st.markdown("""
<style>
/* ── 全体背景 ── */
.stApp {
    background: linear-gradient(160deg, #f0f4ff 0%, #fafafa 60%, #fff8f0 100%);
}

/* ── ヘッダーバナー ── */
.app-header {
    background: linear-gradient(135deg, #5ba3d9 0%, #7ec8e3 55%, #b8e8f7 100%);
    border-radius: 16px;
    padding: 36px 40px;
    margin-bottom: 28px;
    box-shadow: 0 6px 24px rgba(91, 163, 217, 0.28);
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.app-header h1 {
    background: linear-gradient(90deg, #1b4f8a 0%, #3a7ec8 45%, #2ab5a5 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    font-size: 3.8rem;
    font-weight: 800;
    margin: 0;
    letter-spacing: 4px;
    padding-bottom: 14px;
    border-bottom: 2px solid rgba(255,255,255,0.45);
    display: inline-block;
}
.header-icon {
    animation: float 3.5s ease-in-out infinite;
    filter: drop-shadow(0 4px 12px rgba(30,90,150,0.25));
}
@keyframes float {
    0%, 100% { transform: translateY(0px); }
    50%       { transform: translateY(-8px); }
}
.app-header p {
    color: rgba(255,255,255,0.75);
    font-size: 0.9rem;
    margin: 0;
}

/* ── セクションカード ── */
.section-card {
    background: #ffffff;
    border-radius: 12px;
    padding: 22px 26px;
    margin-bottom: 18px;
    border-left: 5px solid #2e6da4;
    box-shadow: 0 2px 10px rgba(0,0,0,0.06);
}
.section-card.gold {
    border-left-color: #c9a227;
}
.section-card.teal {
    border-left-color: #2ab5a5;
}
.section-num {
    display: inline-block;
    background: #2e6da4;
    color: white;
    border-radius: 50%;
    width: 28px;
    height: 28px;
    line-height: 28px;
    text-align: center;
    font-size: 0.85rem;
    font-weight: 700;
    margin-right: 10px;
}
.section-num.gold { background: #c9a227; }
.section-num.teal { background: #2ab5a5; }
.section-title {
    font-size: 1.05rem;
    font-weight: 700;
    color: #1b3a6b;
    vertical-align: middle;
}

/* ── メトリクスカード ── */
[data-testid="stMetric"] {
    background: #ffffff;
    border-radius: 10px;
    padding: 14px 16px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    border-top: 3px solid #2e6da4;
}
[data-testid="stMetricLabel"] { color: #666; font-size: 0.8rem; }
[data-testid="stMetricValue"] { color: #1b3a6b; font-weight: 700; }

/* ── 生成ボタン ── */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1b3a6b, #2e6da4);
    color: white;
    border: none;
    border-radius: 10px;
    padding: 14px 40px;
    font-size: 1.05rem;
    font-weight: 700;
    letter-spacing: 0.5px;
    box-shadow: 0 4px 14px rgba(46,109,164,0.35);
    transition: all 0.2s;
    width: 100%;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #2e6da4, #4eacd1);
    box-shadow: 0 6px 20px rgba(46,109,164,0.45);
    transform: translateY(-1px);
}

/* ── ダウンロードボタン ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #1a7a5e, #2ab5a5);
    color: white;
    border: none;
    border-radius: 10px;
    padding: 12px 32px;
    font-size: 1rem;
    font-weight: 700;
    box-shadow: 0 4px 14px rgba(42,181,165,0.35);
    width: 100%;
}

/* ── 区切り線 ── */
hr { border-color: #dde4f0; }

/* ── テーブル ── */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* ── selectbox ── */
[data-testid="stSelectbox"] > div > div {
    border-radius: 8px;
    border-color: #2e6da4;
}

/* ── アップロードボタン ── */
[data-testid="stFileUploader"] button {
    background: linear-gradient(135deg, #2e6da4, #4eacd1);
    color: white;
    border: none;
    border-radius: 8px;
    font-weight: 700;
}
</style>
""", unsafe_allow_html=True)

PROPERTIES = {
    "MU HOUSE IKEBUKURO": {
        "preparer": "鄭光男",
        "file_prefix": "MU HOUSE IKEBUKURO_収益計算_",
        "sheet_name_template": "MU HOUSE IKEBUKUR - {year}年{month}月_損益",
        "hotel_name": "MU HOUSE IKEBUKURO",
    },
    "Otsuka Stay Base": {
        "preparer": "鄭光男",
        "file_prefix": "Otsuka Stay Base_収益計算_",
        "sheet_name_template": "Otsuka Stay Base - {year}年{month}月_損益",
        "hotel_name": "Otsuka Stay Base",
    },
}

MAX_INCOME_ROWS = 8
MAX_CLEANING_ROWS = 6


def parse_airbnb_csv(file_obj):
    df = pd.read_csv(file_obj, encoding="utf-8-sig")
    df = df[df["種別"] == "予約"].copy()
    df["開始日_dt"] = pd.to_datetime(df["開始日"], format="%m/%d/%Y")
    df["終了日_dt"] = pd.to_datetime(df["終了日"], format="%m/%d/%Y")
    df["サービス料_数値"] = (
        df["サービス料"].astype(str).str.replace(",", "").str.strip().replace("nan", "0").astype(float)
    )
    df["総収入_数値"] = (
        df["総収入"].astype(str).str.replace(",", "").str.strip().astype(float)
    )
    return df.sort_values("開始日_dt").reset_index(drop=True)


def fill_excel(wb, df, year, month, prop, today):
    ws = wb.worksheets[0]
    ws.title = prop["sheet_name_template"].format(year=year, month=month)

    ws["B5"] = prop["hotel_name"]
    ws["F5"] = df["開始日_dt"].min().to_pydatetime()
    ws["G5"] = df["終了日_dt"].max().to_pydatetime()

    today_dt = datetime(today.year, today.month, today.day)
    ws["B8"] = prop["preparer"]
    ws["D8"] = today_dt
    ws["E8"] = prop["preparer"]
    ws["G8"] = today_dt

    for r in range(11, 20):
        ws[f"C{r}"] = None
        ws[f"G{r}"] = None

    for i, (_, row) in enumerate(df.iterrows()):
        if i >= MAX_INCOME_ROWS:
            break
        s, e = row["開始日_dt"], row["終了日_dt"]
        ws[f"C{11 + i}"] = f"OTA(airbnb)ゲスト予約({s.year}/{s.month}/{s.day}-{e.year}/{e.month}/{e.day})"
        ws[f"G{11 + i}"] = int(row["総収入_数値"])

    ws["G26"] = int(df["サービス料_数値"].sum())

    for r in range(27, 35):
        ws[f"C{r}"] = None
        ws[f"G{r}"] = None

    for i, (_, row) in enumerate(df.iterrows()):
        if i >= MAX_CLEANING_ROWS:
            break
        checkout = row["終了日_dt"]
        ws[f"C{27 + i}"] = f"清掃費{checkout.month}/{checkout.day}"
        ws[f"G{27 + i}"] = 13500

    return wb


# ── UI ──────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
  <h1>InnSight</h1>
  <div class="header-icon">
  <svg width="150" height="135" viewBox="0 0 110 100" fill="none" xmlns="http://www.w3.org/2000/svg">
    <!-- 月 -->
    <path d="M72,18 Q44,30 44,52 Q44,74 66,84 Q36,82 24,58 Q14,36 30,20 Q48,6 72,18Z"
          fill="#FFD966"/>
    <!-- 星1（大） -->
    <path d="M88,22 L90,28 L96,28 L91,32 L93,38 L88,34 L83,38 L85,32 L80,28 L86,28 Z"
          fill="white">
      <animate attributeName="opacity" values="1;0.3;1" dur="2.2s" repeatCount="indefinite"/>
    </path>
    <!-- 星2（中） -->
    <path d="M96,56 L97.5,60 L102,60 L98.5,63 L100,67 L96,64.5 L92,67 L93.5,63 L90,60 L94.5,60 Z"
          fill="white" opacity="0.85">
      <animate attributeName="opacity" values="0.3;1;0.3" dur="1.8s" repeatCount="indefinite"/>
    </path>
    <!-- 星3（小） -->
    <path d="M82,76 L83,79 L86,79 L83.5,81 L84.5,84 L82,82.5 L79.5,84 L80.5,81 L78,79 L81,79 Z"
          fill="white" opacity="0.75">
      <animate attributeName="opacity" values="0.6;0.15;0.6" dur="2.8s" repeatCount="indefinite"/>
    </path>
    <!-- 小さな光の粒 -->
    <circle cx="100" cy="38" r="2" fill="white">
      <animate attributeName="opacity" values="0.2;0.9;0.2" dur="3.5s" repeatCount="indefinite"/>
    </circle>
    <circle cx="76" cy="88" r="1.5" fill="white">
      <animate attributeName="opacity" values="0.8;0.1;0.8" dur="2s" repeatCount="indefinite"/>
    </circle>
  </svg>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="section-card">
  <span class="section-num">物</span>
  <span class="section-title">物件を選択</span>
</div>
""", unsafe_allow_html=True)
property_name = st.selectbox("", list(PROPERTIES.keys()), label_visibility="collapsed")
prop_config = PROPERTIES[property_name]

st.markdown("""
<div class="section-card">
  <span class="section-num">1</span>
  <span class="section-title">CSVをアップロード</span>
</div>
""", unsafe_allow_html=True)
csv_files = st.file_uploader(
    "CSVファイルをアップロード（複数可）", type=["csv", "txt"], accept_multiple_files=True
)

st.markdown("""
<div class="section-card gold">
  <span class="section-num gold">2</span>
  <span class="section-title">テンプレートExcelをアップロード</span>
</div>
""", unsafe_allow_html=True)
st.caption("前月などの既存Excelファイルをアップロードしてください。ホテル名・シート名は自動で書き換わります。")
template_file = st.file_uploader("Excelテンプレート", type=["xlsx"])

components.html("""
<script>
function styleUploadButtons() {
    const btns = window.parent.document.querySelectorAll('[data-testid="stFileUploader"] button');
    const styles = [
        'linear-gradient(135deg, #2e6da4, #4eacd1)',
        'linear-gradient(135deg, #c9a227, #f0c84a)'
    ];
    btns.forEach((btn, i) => {
        if (i < styles.length) {
            btn.style.background = styles[i];
            btn.style.color = 'white';
            btn.style.border = 'none';
            btn.style.borderRadius = '8px';
            btn.style.fontWeight = '700';
        }
    });
}
const observer = new MutationObserver(styleUploadButtons);
observer.observe(window.parent.document.body, { childList: true, subtree: true });
styleUploadButtons();
</script>
""", height=0)

st.markdown("""
<div class="section-card teal">
  <span class="section-num teal">3</span>
  <span class="section-title">作成日</span>
</div>
""", unsafe_allow_html=True)
today = st.date_input("", value=date.today(), label_visibility="collapsed")

combined_df = None
year, month = date.today().year, date.today().month

if csv_files:
    all_dfs = []
    for f in csv_files:
        try:
            all_dfs.append(parse_airbnb_csv(f))
        except Exception as e:
            st.error(f"{f.name} の読み込みエラー: {e}")

    if all_dfs:
        combined_df = pd.concat(all_dfs, ignore_index=True).sort_values("開始日_dt").reset_index(drop=True)

        min_date = combined_df["開始日_dt"].min()
        col1, col2 = st.columns(2)
        with col1:
            year = st.number_input("年", value=int(min_date.year), min_value=2020, max_value=2035)
        with col2:
            month = st.number_input("月", value=int(min_date.month), min_value=1, max_value=12)

        if len(combined_df) > MAX_INCOME_ROWS:
            st.warning(f"予約件数が{MAX_INCOME_ROWS}件を超えています（{len(combined_df)}件）。収入明細は最初の{MAX_INCOME_ROWS}件のみ記入されます。")
        if len(combined_df) > MAX_CLEANING_ROWS:
            st.warning(f"清掃費はテンプレートの上限（{MAX_CLEANING_ROWS}件）を超えているため、最初の{MAX_CLEANING_ROWS}件のみ記入されます。")

        st.subheader("データプレビュー")
        preview = combined_df[["開始日_dt", "終了日_dt", "総収入_数値", "サービス料_数値"]].copy()
        preview.columns = ["チェックイン", "チェックアウト", "総収入", "OTAサービス料"]
        preview["チェックイン"] = preview["チェックイン"].dt.strftime("%Y/%m/%d")
        preview["チェックアウト"] = preview["チェックアウト"].dt.strftime("%Y/%m/%d")
        preview["総収入"] = preview["総収入"].apply(lambda x: f"¥{x:,.0f}")
        preview["OTAサービス料"] = preview["OTAサービス料"].apply(lambda x: f"¥{x:,.0f}")
        st.dataframe(preview, use_container_width=True, hide_index=True)

        c1, c2 = st.columns(2)
        c1.metric("予約件数", f"{len(combined_df)}件")
        c2.metric("総収入合計", f"¥{combined_df['総収入_数値'].sum():,.0f}")
        c3, c4 = st.columns(2)
        c3.metric("OTAサービス料", f"¥{combined_df['サービス料_数値'].sum():,.0f}")
        c4.metric("清掃費合計", f"¥{min(len(combined_df), MAX_CLEANING_ROWS) * 13500:,}")

st.divider()

can_generate = combined_df is not None and template_file is not None
if st.button("Excelを生成する", type="primary", disabled=not can_generate):
    try:
        wb = load_workbook(template_file)
        wb = fill_excel(wb, combined_df, int(year), int(month), prop_config, today)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{prop_config['file_prefix']}{int(year)}{int(month):02d}.xlsx"
        st.balloons()
        st.markdown(f"""
<div class="success-banner">
    <div class="success-icon">✓</div>
    <div class="success-text">{filename} が生成されました</div>
</div>
<style>
.success-banner {{
    display: flex;
    align-items: center;
    gap: 16px;
    background: linear-gradient(135deg, #1a7a5e, #2ab5a5);
    border-radius: 12px;
    padding: 20px 28px;
    margin: 12px 0;
    animation: slideIn 0.5s ease-out;
    box-shadow: 0 4px 16px rgba(42,181,165,0.35);
}}
.success-icon {{
    font-size: 2rem;
    font-weight: 900;
    color: white;
    animation: popIn 0.4s 0.3s both;
}}
.success-text {{
    color: white;
    font-size: 1.05rem;
    font-weight: 700;
    animation: fadeIn 0.5s 0.2s both;
}}
@keyframes slideIn {{
    from {{ transform: translateY(-20px); opacity: 0; }}
    to   {{ transform: translateY(0);    opacity: 1; }}
}}
@keyframes popIn {{
    0%   {{ transform: scale(0); }}
    70%  {{ transform: scale(1.3); }}
    100% {{ transform: scale(1); }}
}}
@keyframes fadeIn {{
    from {{ opacity: 0; transform: translateX(-10px); }}
    to   {{ opacity: 1; transform: translateX(0); }}
}}
</style>
""", unsafe_allow_html=True)
        st.download_button(
            label="Excelをダウンロード",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error(f"エラーが発生しました: {e}")
        import traceback
        st.code(traceback.format_exc())
